from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
from copy import copy
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse, Response
from openpyxl import load_workbook
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field


APP_DIR = Path(__file__).resolve().parent
REPO_ROOT = APP_DIR.parent.parent
DEFAULT_TEMPLATE_PATH = (
    REPO_ROOT / "assets" / "templates" / "invoice_table_template.xlsx"
)
TEMPLATE_PATH = Path(
    os.environ.get("INVOICE_TEMPLATE_PATH", str(DEFAULT_TEMPLATE_PATH))
).resolve()
SOFFICE_BINARY = os.environ.get("SOFFICE_BINARY", "soffice")

FIRST_DATA_ROW = 6
SUMMARY_START_ROW = 21
BASE_TEMPLATE_ROWS = SUMMARY_START_ROW - FIRST_DATA_ROW

XL_DOUBLE = "double"
XL_THIN = "thin"

app = FastAPI(title="CVANT Invoice Render Service", version="1.0.0")


class RowPayload(BaseModel):
    no: str = ""
    tanggal: str = ""
    plat: str = ""
    muatan: str = ""
    muat: str = ""
    bongkar: str = ""
    tonase: str = ""
    harga: str = ""
    total: str = ""


class InvoiceRenderPayload(BaseModel):
    rowCount: int = Field(..., ge=1)
    rows: list[RowPayload]
    renderMode: str = Field(default="table")
    summaryValues: dict[str, str] | None = None


def _copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def _copy_row_format(ws, source_row: int, target_row: int) -> None:
    for col in range(1, 10):
        _copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))
    source_height = ws.row_dimensions[source_row].height
    if source_height is not None:
        ws.row_dimensions[target_row].height = source_height


def _set_cell_text(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row, col)
    cell.number_format = "@"
    cell.value = text or ""


def _apply_header_style(ws, last_data_row: int) -> None:
    thin_side = Side(style=XL_THIN, color="000000")
    double_side = Side(style=XL_DOUBLE, color="000000")

    for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=9):
        for cell in row:
            cell.fill = PatternFill(fill_type=None)
            current_font = copy(cell.font) if cell.font else Font()
            cell.font = Font(
                name=current_font.name,
                charset=current_font.charset,
                family=current_font.family,
                b=True,
                i=current_font.i,
                strike=current_font.strike,
                outline=current_font.outline,
                shadow=current_font.shadow,
                condense=current_font.condense,
                color="000000",
                extend=current_font.extend,
                sz=current_font.sz,
                u=current_font.u,
                vertAlign=current_font.vertAlign,
                scheme=current_font.scheme,
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=double_side,
                bottom=double_side,
            )

    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=last_data_row, min_col=1, max_col=9):
        for cell in row:
            cell.fill = PatternFill(fill_type=None)
            current_font = copy(cell.font) if cell.font else Font()
            cell.font = Font(
                name=current_font.name,
                charset=current_font.charset,
                family=current_font.family,
                b=current_font.b,
                i=current_font.i,
                strike=current_font.strike,
                outline=current_font.outline,
                shadow=current_font.shadow,
                condense=current_font.condense,
                color="000000",
                extend=current_font.extend,
                sz=current_font.sz,
                u=current_font.u,
                vertAlign=current_font.vertAlign,
                scheme=current_font.scheme,
            )
            horizontal = "right" if cell.column >= 8 else "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=cell.border.top if cell.border else Side(style=None),
                bottom=thin_side,
            )


def _apply_summary(ws, render_mode: str, actual_summary_row: int, summary_values: dict[str, str] | None) -> None:
    if not summary_values:
        return

    thin_side = Side(style=XL_THIN, color="000000")
    bold_font = Font(bold=True, color="000000")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    if render_mode == "table_with_total":
        _set_cell_text(ws, actual_summary_row, 8, "TOTAL BAYAR Rp.")
        _set_cell_text(ws, actual_summary_row, 9, summary_values.get("total", ""))
        ws.cell(actual_summary_row, 8).font = bold_font
        ws.cell(actual_summary_row, 8).alignment = right_align
        ws.cell(actual_summary_row, 9).font = bold_font
        ws.cell(actual_summary_row, 9).alignment = right_align
        ws.cell(actual_summary_row, 9).border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )
        return

    _set_cell_text(ws, actual_summary_row, 2, "Hormat kami,")
    ws.cell(actual_summary_row, 2).alignment = center_align

    labels = (
        ("SUBTOTAL Rp.", "subtotal"),
        ("PPH 2% Rp.", "pph"),
        ("TOTAL BAYAR Rp.", "total"),
    )
    for offset, (label, key) in enumerate(labels):
        row = actual_summary_row + offset
        _set_cell_text(ws, row, 8, label)
        _set_cell_text(ws, row, 9, summary_values.get(key, ""))
        ws.cell(row, 8).font = bold_font
        ws.cell(row, 8).alignment = right_align
        ws.cell(row, 9).font = bold_font
        ws.cell(row, 9).alignment = right_align
        ws.cell(row, 9).border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )


def _configure_print_area(ws, render_mode: str, last_data_row: int, actual_summary_row: int) -> None:
    if render_mode == "table_with_total":
        ws.print_area = f"A5:I{actual_summary_row}"
    elif render_mode == "table_with_summary":
        ws.print_area = f"A5:I{actual_summary_row + 2}"
    else:
        ws.print_area = f"A5:I{last_data_row}"

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.05
    ws.page_margins.right = 0.05
    ws.page_margins.top = 0.05
    ws.page_margins.bottom = 0.05


def _find_soffice_binary() -> str:
    candidates = [SOFFICE_BINARY]
    if os.name == "nt":
        candidates.extend(
            [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]
        )
    for candidate in candidates:
        if shutil.which(candidate) or Path(candidate).exists():
            return candidate
    raise FileNotFoundError(
        "LibreOffice/soffice tidak ditemukan. Set env SOFFICE_BINARY atau install LibreOffice."
    )


def _convert_xlsx_to_pdf(xlsx_path: Path, output_dir: Path) -> Path:
    soffice = _find_soffice_binary()
    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_dir),
            str(xlsx_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed: {result.stderr or result.stdout}"
        )

    pdf_path = output_dir / f"{xlsx_path.stem}.pdf"
    if not pdf_path.exists():
        raise FileNotFoundError("PDF output was not generated by LibreOffice.")
    return pdf_path


def render_invoice_pdf(payload: InvoiceRenderPayload) -> bytes:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template Excel tidak ditemukan: {TEMPLATE_PATH}")

    with tempfile.TemporaryDirectory(prefix="cvant_invoice_render_") as temp_dir:
        temp_path = Path(temp_dir)
        workbook = load_workbook(TEMPLATE_PATH)
        ws = workbook.active

        row_count = max(1, payload.rowCount)
        extra_rows = max(0, row_count - BASE_TEMPLATE_ROWS)
        if extra_rows:
            for offset in range(extra_rows):
                insert_row = SUMMARY_START_ROW + offset
                ws.insert_rows(insert_row)
                _copy_row_format(ws, 20, insert_row)

        last_data_row = FIRST_DATA_ROW + row_count - 1
        actual_summary_row = SUMMARY_START_ROW + extra_rows

        for row_number in range(FIRST_DATA_ROW, last_data_row + 1):
            for col in range(1, 10):
                _set_cell_text(ws, row_number, col, "")

        rows = payload.rows
        for index in range(row_count):
            row_number = FIRST_DATA_ROW + index
            row = rows[index] if index < len(rows) else RowPayload()
            _set_cell_text(ws, row_number, 1, row.no)
            _set_cell_text(ws, row_number, 2, row.tanggal)
            _set_cell_text(ws, row_number, 3, row.plat)
            _set_cell_text(ws, row_number, 4, row.muatan)
            _set_cell_text(ws, row_number, 5, row.muat)
            _set_cell_text(ws, row_number, 6, row.bongkar)
            _set_cell_text(ws, row_number, 7, row.tonase)
            _set_cell_text(ws, row_number, 8, row.harga)
            _set_cell_text(ws, row_number, 9, row.total)

        _apply_header_style(ws, last_data_row)
        _apply_summary(
            ws,
            render_mode=payload.renderMode,
            actual_summary_row=actual_summary_row,
            summary_values=payload.summaryValues,
        )
        _configure_print_area(
            ws,
            render_mode=payload.renderMode,
            last_data_row=last_data_row,
            actual_summary_row=actual_summary_row,
        )

        xlsx_output = temp_path / "invoice_table.xlsx"
        workbook.save(xlsx_output)
        workbook.close()

        pdf_path = _convert_xlsx_to_pdf(xlsx_output, temp_path)
        return pdf_path.read_bytes()


@app.get("/health")
def health() -> dict[str, Any]:
    return {
        "ok": True,
        "service": "invoice-render-service",
        "templatePath": str(TEMPLATE_PATH),
    }


@app.post("/render-table")
def render_table(payload: InvoiceRenderPayload) -> Response:
    try:
        pdf_bytes = render_invoice_pdf(payload)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return Response(content=pdf_bytes, media_type="application/pdf")


@app.exception_handler(Exception)
async def unhandled_exception_handler(_, exc: Exception) -> JSONResponse:
    return JSONResponse(
        status_code=500,
        content={
            "error": exc.__class__.__name__,
            "message": str(exc),
        },
    )
